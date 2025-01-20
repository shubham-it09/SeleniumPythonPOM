import openpyxl


path="..//excel//testdata.xlsx"
sheetname="LoginTest"



def getrowCount(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row

def getcolCount(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column



def getcellData(path,sheetname,rowNum,colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(rowNum,colNum).value

def setcellData(path,sheetname,rowNum,colNum,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(rowNum,colNum).value=data
    workbook.save(path)


rows = getrowCount(path,sheetname)
cols = getcolCount(path,sheetname)
print(rows ,"----",cols)

# print(getcellData(path,sheetname,2,1))
# setcellData(path,sheetname,1,4,"DOB")